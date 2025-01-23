from pprint import pprint
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from requests import Session

from parse_l2 import authorization_l2
from settings import login_l2, password_l2


class History:

    def __init__(self, connection: Session, history_number: int):
        self.connection = connection
        self.history_number = history_number

        try:
            json_data = {
                'direction': history_number,
                'r_type': 'all',
                'every': False,
            }

            response = self.connection.post(
                'http://192.168.10.161/api/stationar/directions-by-key',
                json=json_data,
                verify=False
            )
            data = response.json().get('data')
            self._examination_numbers = data

        except Exception as e:
            print(f'{e}: ошибка получения номера первичного осмотра при инициализации объекта')
            self._examination_numbers = None

    def _get_first_examination_data(self):
        for record in self._examination_numbers:
            examination_title = record.get('researches')
            if examination_title == ['Первичный осмотр-травматология (при поступлении)']:
                first_examination_number = record.get('pk')
                json_data = {
                    'pk': first_examination_number,
                    'force': True,
                }

                examination_data: dict = self.connection.post(
                    'http://192.168.10.161/api/directions/paraclinic_form',
                    json=json_data,
                    verify=False
                ).json()

                return examination_data

    def _get_diaries_data(self):
        diaries_data = []
        for record in self._examination_numbers:
            if record.get('researches') == ['Осмотр']:
                inspection_number = record.get('pk')
                json_data = {
                    'pk': inspection_number,
                    'force': True,
                }

                examination_data: dict = self.connection.post(
                    'http://192.168.10.161/api/directions/paraclinic_form',
                    json=json_data,
                    verify=False
                ).json()
                examination_date = examination_data.get('researches')[0].get('research').get('groups')[1].get('fields')[0].get('value')
                examination_weekday = examination_data.get('researches')[0].get('research').get('groups')[1].get('fields')[1].get('value')
                examination_type = examination_data.get('researches')[0].get('research').get('groups')[0].get('fields')[0].get('value')
                diaries_data.append({
                    'дата осмотра': examination_date,
                    'день недели': examination_weekday,
                    'тип осмотра': examination_type,
                })
        return diaries_data

    def check_inspection_manager(self):
        """Проверка наличия осмотра заведующим в первые 48 часов (рабочие дни, без учета праздников)"""
        for examination_data in self._examination_numbers:
            if examination_data.get('researches') == ['Первичный осмотр-травматология (при поступлении)']:
                first_examination_date = examination_data.get('date_create').split(' - ')[0]
                first_examination_weekday = examination_data.get('date_create').split(' - ')[1]

                for item in self._get_diaries_data()[::-1]:
                    if item.get('тип осмотра') in ['лечащим врачом совместно с заведующим отделением', 'заведующим отделением']:
                        date = '.'.join(item.get('дата осмотра').split('-')[::-1])
                        examination_date_datetime = datetime.strptime(date, '%d.%m.%Y')
                        first_examination_date_datetime = datetime.strptime(first_examination_date, '%d.%m.%Y')
                        time_delta = (examination_date_datetime - first_examination_date_datetime).days
                        print('Не учитываются праздничные дни, только выходные дни недели')
                        if time_delta < 3:
                            print('Осмотр с заведующим отделением проведен в течение 2х дней')
                            return True
                        elif 5 > time_delta >= 3 and first_examination_weekday == 'ПТ':
                            print('Пациент поступил в пятницу и осмотр с заведующим проведен в течение 2х рабочих дней')
                            return True
                        elif 4 > time_delta >= 3 and first_examination_weekday == 'СБ':
                            print('Пациент поступил в субботу и осмотр с заведующим проведен в течение 2х рабочих дней')
                            return True
                        else:
                            print('Осмотр с заведующим проведен позже 2х рабочих дней')
                        return False

    def check_first_examination(self):
        examination_data = self._get_first_examination_data()

        """Проверка даты и времени поступления и первичного осмотра"""
        date_hospitalization = examination_data.get('researches')[0].get('research').get('groups')[0].get('fields')
        dates = {}
        for field in date_hospitalization:
            dates[field.get('title')] = field.get('value')

        date_time_receipt = datetime.strptime(
            f'{dates.get("Дата поступления")} {dates.get("Время поступления")}',
            '%Y-%m-%d %H:%M'
        )
        date_time_inspection = datetime.strptime(
            f'{dates.get("Дата осмотра")} {dates.get("Время осмотра")}',
            '%Y-%m-%d %H:%M'
        )
        time_delta = (date_time_inspection - date_time_receipt).total_seconds() / 60
        if time_delta <= 120:
            print('Диагноз при поступлении установлен в течение 2х часов')
            answer_1 = '+'
        else:
            print('Внимание!!! Диагноз установлен позже 2х часов')
            answer_1 = '-'

        """Проверка времени назначения обезболивания при поступлении"""
        analgesics = ('Ибупрофен', 'Кетопрофен', 'Метамизол натрия', 'Парацетамол', 'Нимесулид')
        answer_2 = None
        complaints = examination_data.get('researches')[0].get('research').get('groups')[1].get('fields')[0].get('value')

        prescribed_medications = examination_data.get('researches')[0].get('procedure_list')

        if 'боль' in complaints or 'боли' in complaints:
            for drug in prescribed_medications:
                drug_title: str = drug.get('drug').split(' ')[0]

                if drug_title in analgesics:
                    date_start = drug.get('dateStart')
                    time_start = drug.get('timesSelected')
                    for time in time_start:
                        datetime_start = datetime.strptime(f'{date_start} {time}', '%Y-%m-%d %H:%M')
                        time_delta = (datetime_start - date_time_receipt).total_seconds() / 60
                        if 0 < time_delta <= 60:
                            print('Обезболивание назначено вовремя')
                            answer_2 = 3
                        elif time_delta <= 0:
                            print('Проверить время назначения обезболивания')
                            answer_2 = 1
                        else:
                            print('Обезболивание назначено позже')
                            answer_2 = 2

                else:
                    print('Обезболивание не назначено')
                    answer_2 = 0
        return answer_1, answer_2

    def write_in_table(self):
        """Записывает данные проверки в таблицу"""
        first_check = self.check_first_examination()
        painkiller_time = ''
        if first_check[0] == 0:
            painkiller = '-'
        else:
            painkiller = '+'
            if first_check[1] == 1 or first_check[1] == 2:
                painkiller_time = '-'
            elif first_check[1] == 3:
                painkiller_time = '+'

        examination_check = self.check_inspection_manager()
        if examination_check:
            chief_examination = '+'
        else:
            chief_examination = '-'

        # print(first_check)
        # if first_check[0]:
        #     fill_1 = PatternFill(start_color='BFE5A8', end_color='BFE5A8', fill_type='solid')
        # else:
        #     fill_1 = PatternFill(start_color='E6AC89', end_color='E6AC89', fill_type='solid')

        workbook = load_workbook('/Users/aleksejdegtarev/Desktop/Экспертиза (скрипт).xlsx')
        worksheet = workbook.worksheets[0]
        worksheet.append(
            {
                1: f'{self.history_number}',
                2: f'{self._get_first_examination_data().get("patient").get("fio")}',
                3: first_check[0],
                4: painkiller,
                5: painkiller_time,
                6: chief_examination
            }
        )
        workbook.save('/Users/aleksejdegtarev/Desktop/Экспертиза (скрипт).xlsx')
        return worksheet


numbers = [
    3450609,
    3450136,
    3426798,
    3450799,
    3445665,
    3445975,
    3445474,
    3445879,
    3450172,
    3433333,
    3454387,
    3444656,
    3426702,
    3454611,
    3450303,
    3450412,
    3453805,
    3450002,
    3453722,
    3450045,
    3453565,
    3445723,
    3445831,
    3445915,
    3447816,
    3453538,
    3444730,
    3444983,
    3438766,
    3453660,
    3445803,
    3418490,
    3446740,
    3445850,
    3438071,
    3447246,
    3445629
]

session = Session()

authorization_l2(session, login_l2, password_l2)
for item in numbers:
    test = History(session, item)

    test.write_in_table()
