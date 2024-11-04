import json
from pprint import pprint

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

from parse_l2 import authorization_l2
from settings import login_l2, password_l2


def read_xlsx(path: str) -> list:
    """Получение данных из суточного отчета"""
    workbook = load_workbook(path)

    raw_data = workbook.worksheets[0]
    data = []
    for row in raw_data.values:
        if row[4]:
            data.append(row)
    return data


def find_index(data_in_table):
    """Поиск номера колонки в таблице"""
    param = 'Направление'
    column_number = None
    for row in data_in_table:
        for index, item in enumerate(row):
            if param == item:
                column_number = index
    return column_number


def search_diag_id(diag_mkb: str) -> str:
    """Получение id диагноза по коду МКБ"""
    with open('jsonS/diag_id.json', 'r') as file:
        diag_ids = json.load(file)

    for item in diag_ids:
        if item.get('Diag_Code') == diag_mkb:
            return item.get('Diag_id')


def get_case_content(case_number: int) -> dict:
    session = requests.Session()
    authorization_l2(session, login=login_l2, password=password_l2)

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'Content-Type': 'application/json',
        'DNT': '1',
        'Origin': 'http://192.168.10.161',
        'Proxy-Connection': 'keep-alive',
        'Referer': 'http://192.168.10.161/ui/results/descriptive',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    }

    json_data = {
        'pk': case_number,
        'searchMode': 'direction',
        'withoutIssledovaniye': None,
        'year': 2024,
    }

    response = session.post(
        'http://192.168.10.161/api/directions/paraclinic_form',
        headers=headers,
        json=json_data,
        verify=False,
    ).json()
    session.close()
    return response


def create_content_for_ECP(content_l2):
    content = {
        'localstatus': '',
        'treatments': '',
        'recommendations': ''
    }

    content['diagnos_mkb'] = content_l2.get('researches')[0].get('diagnos')

    date = content_l2.get('researches')[0].get('examination_date')
    date = '.'.join(date.split('-')[::-1])
    content['date'] = date

    time_case = content_l2.get('researches')[0].get('research').get('groups')[0].get('fields')[1].get('value')
    content['time'] = time_case

    raw_complaints = content_l2.get('researches')[0].get('research').get('groups')[1].get('fields')
    complaints = ''
    for item in raw_complaints:
        if len(item.get('value')) > 0:
            complaints += f"{item.get('title')} {item.get('value')}"
    content['complaint'] = complaints

    anamnesis_morbi = content_l2.get('researches')[0].get('research').get('groups')[2].get('fields')
    for item in anamnesis_morbi:
        if item.get('title') == 'Вид травмы':
            content['trauma_type'] = item.get('value')
        elif item.get('title') == 'Обстоятельства травмы':
            content['anamnesmorbi'] = item.get('value')

    status_localis = content_l2.get('researches')[0].get('research').get('groups')[3].get('fields')
    for item in status_localis:
        if len(item.get('value')) > 0:
            if item.get('title') != 'Обследование':
                content['localstatus'] += f"{item.get('title')}: {item.get('value')}\n"
            else:
                content['research'] = item.get('value')

    diagnosis_text = content_l2.get('researches')[0].get('research').get('groups')[4].get('fields')[0]
    content['diagnos'] = diagnosis_text.get('value')

    treatment = content_l2.get('researches')[0].get('research').get('groups')[5].get('fields')[0]
    if len(treatment.get('value')) > 0:
        content['treatment'] = treatment.get('value')

    recommendations = content_l2.get('researches')[0].get('research').get('groups')[6].get('fields')
    for item in recommendations:
        if len(item.get('value')) > 0:
            content['recommendations'] += f"{item.get('title')}: {item.get('value')}"
    return content


def read_xlsx_ECP_table(path: str):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.worksheets[0]
    return worksheet


def edit_xlsx_ECP_table(path: str, index: str, value: int):
    workbook = load_workbook(path)
    worksheet = workbook.worksheets[0]
    cell = worksheet[index]
    cell.value = value
    cell.alignment = Alignment(horizontal='center', vertical='center')
    workbook.save(path)
    return f'Данные ячейки {index} обновлены'


def cell_color_edit(path: str, cell_number: str):
    workbook = load_workbook(path)
    worksheet = workbook.worksheets[0]
    cell = worksheet[cell_number]
    cell_department = worksheet[cell_number.replace('AJ', 'A')]
    ecp_data = cell.value.split('/')[0].lstrip('=')
    fact_data = cell.value.split('/')[1]
    ecp_cell = worksheet[ecp_data]
    fact_cell = worksheet[fact_data]

    cell_percent = ecp_cell.value / fact_cell.value * 100
    if cell_percent > 80:
        fill = PatternFill(start_color='BFE5A8', end_color='BFE5A8', fill_type='solid') # BFE5A8 - green, ADC8E9 - blue, E6AC89 - orange
    elif 50 < cell_percent < 80:
        fill = PatternFill(start_color='ADC8E9', end_color='ADC8E9', fill_type='solid')
    else:
        fill = PatternFill(start_color='E6AC89', end_color='E6AC89', fill_type='solid')
    cell.fill = fill
    cell_department.fill = fill
    workbook.save(path)


def read_xlsx_history(path: str) -> list:
    """Получение данных из суточного отчета"""
    workbook = load_workbook(path)
    data = []
    # for sheet in workbook.worksheets[:4]:
    #     for row in sheet.values:
    #         if row[0] is None or 'Медицинская карта стационарного больного' in row[0]:
    #             pass
    #         else:
    #             split_list = row[0].split(' ')
    #             last_element = split_list[-1].lstrip('№')
    #             data.append(last_element)
    # return data
    for row in workbook.worksheets[2].values:
        if row[0] is None or 'Медицинская карта стационарного больного' in row[0]:
            pass
        else:
            split_list = row[0].split(' ')
            last_element = split_list[-1].lstrip('№')
            data.append(last_element)
    return data


# print(read_xlsx_history('/Users/aleksejdegtarev/PycharmProjects/integration_with_single_digital_platform/ЖУРНАЛ ЭО II уровня 2024.xlsx'))
